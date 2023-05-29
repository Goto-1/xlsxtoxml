import openpyxl
import xml.etree.ElementTree as ET

# Define variable to load the wookbook
wookbook = openpyxl.load_workbook(input('Введите имя контактной карты, без расширения \n')+'.xlsx')
#d = {'Бованенковское месторождение':'БВН','Восточно-Мессояхское месторождение':'МЕС','Находкинское месторождение':'Нхд','Опорная база г. Коротчаево':'Кор','Опорная база с. Газ-Сале':'Газ','Офис г. Москва':'МСК','Офис г. Тюмень':'ТМН','Тазовское месторождение':'ТАЗ','Харасавэйское месторождение':'ХРСВ','Центр Аудио-Конференций':'АКС'}

# Define variable to read the active sheet:
worksheet = wookbook["ОБЩЕЕ"]
l = []
#cycle for tum
for row in range(37, worksheet.max_row):
    lst = []
    lst.append(worksheet[row][0].value)
    lst.append(worksheet[row][1].value)
    l.append(lst)

#cycle for msk
for row in range(37, worksheet.max_row):
    lst = []
    lst.append(worksheet[row][7].value)
    lst.append(worksheet[row][8].value)
    l.append(lst)

#cycle for taz
for row in range(37, worksheet.max_row):
    lst = []
    lst.append(worksheet[row][10].value)
    lst.append(worksheet[row][11].value)
    l.append(lst)

#cycle for mess
for row in range(37, worksheet.max_row):
    lst = []
    lst.append(worksheet[row][16].value)
    lst.append(worksheet[row][17].value)
    l.append(lst)

#cycle for korotchaevo
for row in range(37, worksheet.max_row):
    lst = []
    lst.append(worksheet[row][22].value)
    lst.append(worksheet[row][23].value)
    l.append(lst)

#cycle for bvn
for row in range(37, worksheet.max_row):
    lst = []
    lst.append(worksheet[row][25].value)
    lst.append(worksheet[row][26].value)
    l.append(lst)

#cycle for hrsv
for row in range(37, worksheet.max_row):
    lst = []
    lst.append(worksheet[row][31].value)
    lst.append(worksheet[row][32].value)
    l.append(lst)

print(*l, sep='\n')





XXXIPPhoneDirectory = ET.Element('XXXIPPhoneDirectory', clearlight="true")

title = ET.SubElement(XXXIPPhoneDirectory,'Title')
title.text = 'PSD_PhoneBook'
Prompt = ET.SubElement(XXXIPPhoneDirectory,'Prompt')
Prompt.text = 'Prompt'
for row in range(1, len(l)):
    if l[row][0]:
        if l[row][1] != 'резерв':
            DirectoryEntry = ET.Element('DirectoryEntry')
            Name = ET.SubElement(DirectoryEntry,'Name')
            Name.text = str(l[row][0]) + '|' + l[row][1] 
            Telephone = ET.SubElement(DirectoryEntry,'Telephone')
            Telephone.text = str(l[row][0])
            XXXIPPhoneDirectory.append(DirectoryEntry)

mydata = ET.tostring(XXXIPPhoneDirectory, encoding='utf-8', method='xml')
myfile = open('ContactAll.xml', 'wb')
myfile.write(mydata)

